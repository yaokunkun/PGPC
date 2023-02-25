import math
from random import random, randint

from openpyxl import load_workbook
from typing import List

from Drought import Drought

RUE_max = 3  # 光合作用的最大能量利用率
LAM = 0.62  # 能进行光合作用的叶子部分，占整个植株生物量的百分比
n = 4  # 群落中的物种数 TODO:输入
num_t = 3  # 要研究多少天 TODO:输入
# PWP = 40  # 永久枯萎点(mm)
# WHC = 160  # 土壤蓄水能力（量）(mm)
PARs = [8, 9, 6, 9]  # t天的光照指标 TODO:输入
SLAs = [0.02, 0.03, 0.03, 0.04]  # i种植物的叶面积指标 TODO:输入
PETs = [0.5, 0.5, 0.6]  # t天植物的蒸腾阈值（假设所有植物都一样）TODO:输入
P = [0, 0, 0, 0]  # t天的降水量(mm) TODO:输入
T = [20, 25, 30, 25]  # t天的气温 TODO:输入
ST = [0 for i in range(len(T))]
WR = [0.0 for k in range(num_t + 1)]  # t天的j蓄水量(mm)，需要迭代计算
WR[0] = 0  # TODO:输入
WR_min = 0.15  # 当储水量小于这个值时，开始取地下水
WR_max = 2  # 土壤最大储水量
GW = [0.0 for k in range(num_t + 1)]  # t天的j地下水量(mm)，需要迭代计算
GW[0] = 5  # TODO:输入
GW_max = 2  # 地下水最大储量(mm)
ground_off = 0.05  # 土壤水分流入地下比率
SEN_min = 1  # 生长周期最小值
SEN_max = 3  # 生长周期最大值
start_drought_day = 9999999  # 干旱开始日期（初始化为0，但以实际为准）

alpha = 0.6  # 消光系数（假设所有物种都相同）
alpha_2 = 0.5  # 计算蓄水量WR的时候用到的系数
omega = 45  # 蓄水量影响系数
phi_1 = 775  # 生命周期温度阈值1
phi_2 = 3000  # 生命周期温度阈值2
miu = [0.03, 0.04, 0.05, 0.02]  # 每个植株i的生长周期系数 TODO:输入
# 光合作用相关温度阈值
T_0 = 3
T_1 = [10, 12, 9, 11]  # TODO:输入
T_2 = 20
T_3 = 35
# 植株生物量Biomass  （每天，每个物种）

B = [[0 for i in range(num_t + 1)] for j in range(n)]
root = [0 for i in range(n)]  # TODO:输入  0表示根长小于20cm的物种，1表示根长大于20cm的物种，2表示会分泌激素的物种

long_root_rate = 0  # 长根物种占总物种比例
B[0][0] = 300
B[1][0] = 200
B[2][0] = 250
B[3][0] = 150

dieDate = [-1] * n


def gr_i(i, t):
    PAR = PARs[t]
    gr_i = Gr_i(i, t, PAR) * Wred(t) * Tred(i, t)
    return gr_i


def Gr_i(i, t, PAR):
    """
    第t时间（天），植物i，通过光合作用PAR的生长量
    :param i:
    :param t:
    :param B_i:
    :param PAR:
    :return:
    """
    LAI_i = get_LAI_i(i, t)  # 植物i的光合指标
    LAI_tot = get_LAI_tot(t)
    percent_i = LAI_i / LAI_tot
    Gr_i = 10.0 * PAR * RUE_max * (1 - math.exp(-1.0 * alpha * LAI_tot)) * percent_i
    return Gr_i


def get_LAI_i(i, t):
    """
    得到植物i的光合指标
    :param B_i:
    :return:
    """
    SLA_i = SLAs[i]
    B_i = B[i][t]
    LAI_i = SLA_i * B_i * LAM / 10.0
    return LAI_i


def get_LAI_tot(t):
    LAI_tot = 0
    for i in range(n):
        LAI_tot += get_LAI_i(i, t)
    return LAI_tot


def diff_WR(t):
    """
    计算第t天与第t-1天的蓄水量差值
    :param i:
    :param t:
    :return:
    AET：代表当天植物吸收的土壤储水
    """
    return P[t] - AET(t) - diff_GW(t) + interaction(t)  # 土壤储水量=当天降水量-当天植株吸水量-地下渗水量+（干旱时）根系吸水量


def diff_GW(t):
    if WR[t] > WR_min:
        return WR[t] * ground_off
    else:
        return 0


def AET(t):
    """当天植株用水量"""
    return min(WR[t], ATr(t) + AEv(t))


def ATr(t):
    ATr = PETs[t] * min(1.0, get_LAI_tot(t) / 3) * Wred(t)
    return ATr


def AEv(t):
    AEv = alpha_2 * PETs[t] * (1 - min(1.0, get_LAI_tot(t) / 3))
    return AEv


def Wred(t):
    WR_t = WR[t]
    Wred = 2 / (1 + math.exp(-1.0 * WR_t / omega)) - 1
    return Wred


def Tred(i, t):
    if T[t] <= T_0:
        return 0
    elif T_0 < T[t] < T_1[i]:
        return (T[t] - T_0) / (T_1[i] - T_0)
    elif T_1[i] <= T[t] <= T_2:
        return 1
    elif T_2 <= T[t] <= T_3:
        return (T_3 - T[t]) / (T_3 - T_2)
    else:
        return 0


def SEN(t):
    """计算叶子生长周期"""
    if ST[t] <= phi_1:
        return SEN_min
    elif ST[t] >= phi_2:
        return SEN_max
    else:
        return SEN_min + (SEN_max - SEN_min) * (ST[t] - phi_1) / (phi_2 - phi_1)


def set_ST():
    """算出累积温度数组ST[]"""
    for i in range(len(T)):
        if i == 0:
            ST[i] = T[i]
        elif T[i] > 0:
            ST[i] = ST[i - 1] + T[i]
        else:
            ST[i] = ST[i - 1]


def update_biomass():
    """根据差分方程，更新每种植物每一条的Biomass值于数组B中"""
    for t in range(num_t):
        WR[t + 1] = min(WR[t] + diff_WR(t), WR_max)
        GW[t + 1] = min(GW[t] + diff_GW(t) - interaction(t), GW_max)
        for i in range(n):
            if root[i] != 3:
                B[i][t + 1] = B[i][t] + gr_i(i, t) - miu[i] * SEN(t)
                if WR[t] < WR_min and t - start_drought_day > 7:  # 如果干旱条件下保持了7天
                    root[i] = 3  # 物种i死亡
                    global dieDate
                    dieDate[i] = t
            else:
                B[i][t + 1] = B[i][t]


def interaction(t):
    """当干旱时（即第t天土壤水分=0时），第t天群落中物种之间互惠互利与相互竞争后，带来的土壤储水增量"""
    global start_drought_day
    if WR[t] > WR_min:
        return 0
    else:
        if t > 0 and WR[t - 1] > WR_min or t == 0:
            start_drought_day = t
        return min(GW[t] / (t - start_drought_day + 1) * long_root_rate,
                   PETs[t] * min(1.0, get_LAI_tot(t) / 3))  # 不是AET(t)


def run():
    # 统计群落中各物种根系情况
    num_long_root = 0
    for i in range(n):
        if root[i] == 2:
            num_long_root = n  # 激素使得所有物种均为长根
            break
        elif root[i] == 1:
            num_long_root += 1
    global long_root_rate
    long_root_rate = num_long_root / n

    set_ST()
    update_biomass()
    return B


def main(_n: int, _num_t: int, _PARs: List[float], _SLAs: List[float], _PETs: List[float], _P: List[float],
         _T: List[float], _WR_0: float, _GW_0: float, _miu: List[float], _T_1: List[float], _root: List[int]):
    """

    :param _n: 群落中的物种数
    :param _num_t: 要研究的天数
    :param _PARs:  t天之中的光照指数
    :param _SLAs: i种物种的叶面积指数
    :param _PETs:  t天之中的蒸腾阈值
    :param _P:  t天之中的降水量（mm）
    :param _T:  t天之中的温度
    :param _WR_0:  土壤含水量的初始值
    :param _GW_0:  地下水量的初始值
    :param _miu:  i种植物的生长周期系数
    :param _T_1:  i种植物的光合温度阈值
    :param _root:  i种植物的根系种类
    :return:
    """

    global n
    global num_t
    global PARs
    global SLAs
    global PETs
    global P
    global T
    global WR_0
    global GW_0
    global miu
    global T_1
    global root
    global ST
    global WR
    global GW
    global B
    n = _n
    num_t = _num_t
    PARs = _PARs
    SLAs = _SLAs
    PETs = _PETs
    P = _P
    T = _T
    WR_0 = _WR_0
    GW_0 = _GW_0
    miu = _miu
    T_1 = _T_1
    root = _root

    ST = [0 for i in range(len(T))]
    WR = [0.0 for k in range(num_t + 1)]  # t天的j蓄水量(mm)，需要迭代计算
    GW = [0.0 for k in range(num_t + 1)]  # t天的j地下水量(mm)，需要迭代计算
    WR[0] = WR_0
    GW[0] = GW_0

    B = [[0.0 for i in range(num_t + 1)] for j in range(n)]

    # 初始化B[i][0]
    for i in range(n):
        B[i][0] = 300

    run()


frequency = 3
duration_ave = 80
duration_diff = 15
T_ave = 8
T_diff = 1
T_drought = 20
P, T = Drought(frequency, duration_ave, duration_diff, 3, 1, T_ave, T_diff, T_drought)

PAR = [3.496 * math.exp(0.0605 * T_) + random() * 2.41 for T_ in T]
PET = [0.0844 * math.exp(0.0913 * T_) + random() * 0.84 for T_ in T]

# 读取Excel表格中的数据
row_data = load_workbook(filename='../Plants.xlsx')
sheet = row_data.active
data = []
for column in sheet.iter_cols(values_only=True):
    data.append(column)

SLA = list(data[2])
SLA.remove('SLA')

miu = list(data[3])
miu.remove('miu')

T_1 = list(data[4])
T_1.remove('T')

root_type = list(data[7])
root_type.remove('root type')

num_t = len(T)
n = len(SLA)

main(n, num_t, PAR, SLA, PET, P, T, 1, 1, miu, T_1, root_type)
